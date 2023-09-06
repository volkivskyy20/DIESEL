import openpyxl
import pandas as pd
import pyxlsb
from pyxlsb import open_workbook

print("Please download Oil_Bulletin_Prices_History.xlsx from your inbox folder")
print("Please save the file in XLSB format")
user_input = input("Press any key to start....... ")
def prepare():
    # Load the xlsb file using pyxlsb
    with pyxlsb.open_workbook('Oil_Bulletin_Prices_History.xlsb') as wb:
        with wb.get_sheet(1) as sheet:
            # Create a new workbook using openpyxl
            new_wb = openpyxl.Workbook()
            new_sheet = new_wb.active

            # Copy data from xlsb to xlsx
            for row in sheet.rows():
                new_sheet.append([item.v for item in row])

            # Set column width
            new_sheet.column_dimensions['A'].width = 10.43

            # Save the modified workbook as xlsx
            new_wb.save('Oil_Bulletin_Prices_History_modified.xlsx')
            # Load the Excel file
            wb = openpyxl.load_workbook('Oil_Bulletin_Prices_History_modified.xlsx')

            # Select column A and delete it, shifting remaining columns to the left
            sheet = wb.active
            sheet.delete_cols(1, 1)

            # Delete rows 1 to 6
            sheet.delete_rows(1, 6)

            # Save the modified workbook
            wb.save('Oil_Bulletin_Prices_History_modified.xlsx')


# Call the prepare function
prepare()

print("Please open Oil_Bulletin_Prices_History_modified.xlsx and save it in XLSB format ")
user_input = input("Press '1' to continue... ")

while user_input != '1':
    user_input = input("Invalid input. Press '1' to continue... ")

print("Prossessing")


def transform_data():
    # Replace 'Oil_Bulletin_Prices_History.xlsb' with the actual file path
    input_file_path = 'Oil_Bulletin_Prices_History_modified.xlsb'
    output_file_path = 'output.xlsx'

    # Create a list to hold the transposed data
    transposed_data = []

    # Open the xlsb file
    with open_workbook(input_file_path) as wb:
        # Select the first sheet
        with wb.get_sheet(1) as sheet:
            # Initialize variables to track the current table
            current_table = []
            in_table = False

            # Iterate through rows and columns
            for row in sheet.rows():
                row_data = [item.v for item in row]

                # Check if we're entering or leaving a table
                if all(cell is None for cell in row_data):
                    if in_table:
                        transposed_data.append(current_table)
                        current_table = []
                        in_table = False
                else:
                    current_table.append(row_data)
                    in_table = True

            # If we're still in a table at the end, append it
            if in_table:
                transposed_data.append(current_table)

    # Transpose each table and add it to a list
    transposed_tables = []
    for table in transposed_data:
        df_table = pd.DataFrame(table).transpose()
        transposed_tables.append(df_table)

    # Create the ExcelWriter object
    writer = pd.ExcelWriter(output_file_path, engine='xlsxwriter')

    # Create a Pandas Excel writer using XlsxWriter as the engine.
    writer = pd.ExcelWriter(output_file_path, engine='xlsxwriter')

    # Iterate over transposed tables and save each one as a separate sheet
    for idx, df_table in enumerate(transposed_tables):
        sheet_name = f'TransposedTable_{idx}'

        # Remove index from the DataFrame
        df_table.reset_index(drop=True, inplace=True)

        df_table.to_excel(writer, sheet_name=sheet_name, index=False)

    # Combine all sheets into a single sheet
    combined_df = pd.concat(transposed_tables, ignore_index=True)
    combined_df.to_excel(writer, sheet_name='Combined', index=False)

    # Save and close the ExcelWriter
    writer.close()


transform_data()


def raw():
    # Load the workbook
    workbook = openpyxl.load_workbook('output.xlsx')

    # Select the active sheet
    sheet = workbook['Combined']

    # Convert the generator object to a list and reverse it
    rows = list(sheet.iter_rows(min_row=3, min_col=1, values_only=True))
    rows.reverse()

    # Create a new list to store the filtered rows
    filtered_rows = []

    # Iterate over the reversed rows
    for row in rows:
        if row[0] == ' Gas oil automobile Automotive gas oil Dieselkraftstoff (I)':
            filtered_rows.append(row)

    # Clear the existing sheet
    sheet.delete_rows(3, sheet.max_row)

    # Write the filtered rows back to the sheet
    for row in filtered_rows:
        sheet.append(row)

    sheet.delete_rows(1)
    sheet.delete_cols(2)
    sheet.delete_cols(1)
    for cell in sheet[1]:
        cell.number_format = 'mm/dd/yyyy'
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            # Check if cell has value
            if cell.value:
                # Replace commas in cell value
                if isinstance(cell.value, str):
                    cell.value = cell.value.replace(',', '')
                    cell.value = cell.value.replace('.', ',')
    # Save the modified workbook
    workbook.save('modified_uotput.xlsx')

    # Open the source file
    source_file = "modified_uotput.xlsx"
    source_sheet = "Combined"
    wb_source = openpyxl.load_workbook(source_file)
    ws_source = wb_source[source_sheet]

    # Open the destination file
    destination_file = "Rhenus_Diesel_Floater_.xlsx"
    destination_sheet = "raw data"
    wb_destination = openpyxl.load_workbook(destination_file)
    ws_destination = wb_destination[destination_sheet]

    # Get the range of cells to copy from the source sheet
    start_row = 1
    start_column = 1
    end_row = ws_source.max_row
    end_column = ws_source.max_column

    # Copy the data from the source sheet to the destination sheet
    for row in range(start_row, end_row + 1):
        for column in range(start_column, end_column + 1):
            value = ws_source.cell(row=row, column=column).value
            ws_destination.cell(row=row - start_row + 1, column=column + 3).value = value

    # Save the destination file
    wb_destination.save(destination_file)

raw()

print("Diesel Floater task is finished. The file you need is called Rhenus_Diesel_Floater_.xlsx Please copy paste the "
      "file to another directory. Original file should be left where it is")
user_input = input("Press any to close the program... ")


